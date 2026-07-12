#!/usr/bin/env python3
import argparse
import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.datetime import to_excel


DEFAULT_WORKBOOK = Path(
    "/Users/marcel/Library/CloudStorage/OneDrive-Personal/Documents/Personal/Kitakana/outputs/kitakana_elo_tracker/Kitakana_Elo_Tracker.xlsx"
)
DEFAULT_OUTPUT = Path(__file__).resolve().parents[1] / "supabase" / "seed" / "kitakana_elo_seed.json"


def number(value):
    return None if value in (None, "") else float(value)


def excel_order(value, fallback):
    if value in (None, ""):
        return int(fallback)
    if hasattr(value, "year"):
        return int(to_excel(value))
    return int(value)


def export_seed(workbook_path):
    workbook = load_workbook(workbook_path, data_only=True)
    settings = workbook["Setting"]
    teams_sheet = workbook["Teams"]
    bonuses_sheet = workbook["Bonuses"]
    matches_sheet = workbook["Matches"]

    teams = []
    for row in range(3, teams_sheet.max_row + 1):
        name = teams_sheet.cell(row, 4).value
        if not name:
            continue
        teams.append(
            {
                "name": str(name).strip(),
                "code": str(teams_sheet.cell(row, 5).value or "").strip(),
                "continent": str(teams_sheet.cell(row, 6).value or "").strip(),
                "starting_elo": number(teams_sheet.cell(row, 7).value),
                "expected_current_elo": number(teams_sheet.cell(row, 10).value),
            }
        )
    canonical_names = {team["name"].lower(): team["name"] for team in teams}

    def canonical_team(value):
        raw = str(value or "").strip()
        canonical = canonical_names.get(raw.lower())
        if not canonical:
            raise RuntimeError(f"Team is used in history but missing from Teams: {raw}")
        return canonical

    bonuses = []
    for row in range(3, bonuses_sheet.max_row + 1):
        team = bonuses_sheet.cell(row, 3).value
        if not team:
            continue
        bonus_id = int(bonuses_sheet.cell(row, 1).value)
        bonuses.append(
            {
                "bonus_id": bonus_id,
                "bonus_order": excel_order(bonuses_sheet.cell(row, 2).value, bonus_id),
                "team": canonical_team(team),
                "category": str(bonuses_sheet.cell(row, 4).value or "").strip(),
                "points": number(bonuses_sheet.cell(row, 6).value) or 0,
                "event": str(bonuses_sheet.cell(row, 7).value or "").strip(),
            }
        )

    matches = []
    for row in range(3, matches_sheet.max_row + 1):
        team_a = matches_sheet.cell(row, 3).value
        team_b = matches_sheet.cell(row, 4).value
        if not team_a or not team_b:
            continue
        match_id = int(matches_sheet.cell(row, 1).value)
        winner = str(matches_sheet.cell(row, 5).value or "").strip()
        normalized_winner = "Tie" if winner in {"Tie", "Renga"} else winner
        matches.append(
            {
                "match_code": f"excel-{match_id}",
                # Match ID is the workbook's authoritative sequence. Excel's 1900
                # date-system quirk makes date serials 59 and 60 ambiguous.
                "match_order": match_id,
                "source_match_id": match_id,
                "team_a": canonical_team(team_a),
                "team_b": canonical_team(team_b),
                "winner": normalized_winner,
                "result_type": str(matches_sheet.cell(row, 6).value or "").strip(),
                "tier": str(matches_sheet.cell(row, 7).value or "").strip(),
                "event": str(matches_sheet.cell(row, 21).value or "").strip(),
                "notes": str(matches_sheet.cell(row, 22).value or "").strip(),
                "expected": {
                    "team_a_pre_elo": number(matches_sheet.cell(row, 8).value),
                    "team_b_pre_elo": number(matches_sheet.cell(row, 9).value),
                    "expected_a": number(matches_sheet.cell(row, 10).value),
                    "expected_b": number(matches_sheet.cell(row, 11).value),
                    "result_value": number(matches_sheet.cell(row, 12).value),
                    "multiplier": number(matches_sheet.cell(row, 13).value),
                    "actual_a": number(matches_sheet.cell(row, 14).value),
                    "actual_b": number(matches_sheet.cell(row, 15).value),
                    "team_a_delta": number(matches_sheet.cell(row, 16).value),
                    "team_b_delta": number(matches_sheet.cell(row, 17).value),
                    "team_a_post_elo": number(matches_sheet.cell(row, 18).value),
                    "team_b_post_elo": number(matches_sheet.cell(row, 19).value),
                },
            }
        )

    return {
        "schema": "kitakana-elo-seed",
        "version": 1,
        "source": Path(workbook_path).name,
        "settings": {
            "rating_scale": number(settings["B9"].value),
            "maximum_result_value": number(settings["B10"].value),
        },
        "result_types": {
            str(settings.cell(row, 1).value): number(settings.cell(row, 2).value)
            for row in range(3, 7)
            if settings.cell(row, 1).value
        },
        "tiers": {
            str(settings.cell(row, 5).value): number(settings.cell(row, 6).value)
            for row in range(3, 8)
            if settings.cell(row, 5).value
        },
        "teams": teams,
        "bonuses": bonuses,
        "matches": matches,
    }


def main():
    parser = argparse.ArgumentParser(description="Export the Kitakana Elo workbook into a Supabase seed.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    seed = export_seed(args.workbook.expanduser())
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(seed, indent=2, ensure_ascii=True) + "\n")
    print(
        json.dumps(
            {
                "output": str(args.output),
                "teams": len(seed["teams"]),
                "bonuses": len(seed["bonuses"]),
                "matches": len(seed["matches"]),
            }
        )
    )


if __name__ == "__main__":
    main()
