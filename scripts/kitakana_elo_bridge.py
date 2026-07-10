#!/usr/bin/env python3
import copy
import json
import os
import re
import shutil
import sys
import tempfile
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = Path(
    "/Users/marcel/Library/CloudStorage/OneDrive-Personal/Documents/Personal/Kitakana/outputs/kitakana_elo_tracker/Kitakana_Elo_Tracker.xlsx"
)
DEFAULT_DB = REPO_ROOT / ".kitakana-elo-sync.json"
MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
NS = {"m": MAIN_NS, "r": REL_NS, "pr": PKG_REL_NS}
MATCH_CODE_RE = re.compile(r"\[TCODE:([^\]]+)\]")

ET.register_namespace("", MAIN_NS)
ET.register_namespace("r", REL_NS)


def now_iso():
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def json_out(payload):
    print(json.dumps(payload, default=str))


def workbook_path():
    return Path(os.environ.get("KITAKANA_ELO_WORKBOOK", DEFAULT_WORKBOOK)).expanduser()


def db_path():
    return Path(os.environ.get("KITAKANA_ELO_SYNC_DB", DEFAULT_DB)).expanduser()


def load_db(path):
    if not path.exists():
        return {"version": 1, "matches": {}, "backups": {}}
    try:
        data = json.loads(path.read_text())
        if data.get("version") != 1:
            return {"version": 1, "matches": {}, "backups": {}}
        data.setdefault("matches", {})
        data.setdefault("backups", {})
        return data
    except Exception:
        return {"version": 1, "matches": {}, "backups": {}}


def save_db(path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(data, indent=2, default=str))
    tmp.replace(path)


def make_backup(path, db):
    key = str(path)
    if db.get("backups", {}).get(key):
        return db["backups"][key]
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup = path.with_name(f"{path.stem}.backup-before-tourney-{stamp}{path.suffix}")
    shutil.copy2(path, backup)
    db.setdefault("backups", {})[key] = str(backup)
    return str(backup)


def normalize_space(value):
    return str(value or "").strip()


def excel_cell_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value)


def find_workbook_sheet_path(zip_file, sheet_name):
    workbook = ET.fromstring(zip_file.read("xl/workbook.xml"))
    rels = ET.fromstring(zip_file.read("xl/_rels/workbook.xml.rels"))
    rid = None
    for sheet in workbook.findall(".//m:sheet", NS):
        if sheet.attrib.get("name") == sheet_name:
            rid = sheet.attrib.get(f"{{{REL_NS}}}id")
            break
    if not rid:
        raise RuntimeError(f"Sheet not found: {sheet_name}")
    target = None
    for rel in rels.findall("pr:Relationship", NS):
        if rel.attrib.get("Id") == rid:
            target = rel.attrib.get("Target")
            break
    if not target:
        raise RuntimeError(f"Relationship not found for sheet: {sheet_name}")
    if target.startswith("/"):
        return target.lstrip("/")
    if not target.startswith("xl/"):
        return f"xl/{target}"
    return target


def find_matches_table_path(zip_file):
    for name in zip_file.namelist():
        if not name.startswith("xl/tables/") or not name.endswith(".xml"):
            continue
        text = zip_file.read(name).decode("utf-8", errors="ignore")
        if 'displayName="MatchesTable"' in text or 'name="MatchesTable"' in text:
            return name
    raise RuntimeError("MatchesTable XML not found")


def cell_col(ref):
    return re.sub(r"\d+", "", ref)


def make_cell(ref, style_from=None):
    cell = ET.Element(f"{{{MAIN_NS}}}c", {"r": ref})
    if style_from is not None and style_from.attrib.get("s"):
        cell.attrib["s"] = style_from.attrib["s"]
    return cell


class XlsxXmlEditor:
    def __init__(self, path):
        self.path = Path(path)
        with zipfile.ZipFile(self.path, "r") as zin:
            self.files = {name: zin.read(name) for name in zin.namelist()}
            self.sheet_path = find_workbook_sheet_path(zin, "Matches")
            self.table_path = find_matches_table_path(zin)
        self.sheet_root = ET.fromstring(self.files[self.sheet_path])
        self.table_root = ET.fromstring(self.files[self.table_path])
        self.sheet_data = self.sheet_root.find("m:sheetData", NS)
        self.formula_wb = load_workbook(self.path, data_only=False, keep_links=True)
        self.formula_ws = self.formula_wb["Matches"]

    def table_bounds(self):
        ref = self.table_root.attrib["ref"]
        return range_boundaries(ref)

    def table_end_row(self):
        return self.table_bounds()[3]

    def row_elements(self):
        return {
            int(row.attrib["r"]): row
            for row in self.sheet_data.findall("m:row", NS)
            if row.attrib.get("r", "").isdigit()
        }

    def row(self, row_idx):
        rows = self.row_elements()
        if row_idx in rows:
            return rows[row_idx]
        new_row = ET.Element(f"{{{MAIN_NS}}}row", {"r": str(row_idx), "spans": "1:22"})
        inserted = False
        for idx, row in enumerate(list(self.sheet_data)):
            current = int(row.attrib.get("r", "0") or 0)
            if current > row_idx:
                self.sheet_data.insert(idx, new_row)
                inserted = True
                break
        if not inserted:
            self.sheet_data.append(new_row)
        return new_row

    def find_cell(self, row_el, col):
        ref = f"{col}{row_el.attrib['r']}"
        for cell in row_el.findall("m:c", NS):
            if cell.attrib.get("r") == ref:
                return cell
        return None

    def style_source(self, row_idx, col):
        for source_row in (row_idx, row_idx - 1, 3):
            if source_row < 1:
                continue
            row_el = self.row_elements().get(source_row)
            if not row_el:
                continue
            cell = self.find_cell(row_el, col)
            if cell is not None:
                return cell
        return None

    def get_or_create_cell(self, row_idx, col):
        row_el = self.row(row_idx)
        ref = f"{col}{row_idx}"
        existing = self.find_cell(row_el, col)
        if existing is not None:
            return existing
        new_cell = make_cell(ref, self.style_source(row_idx, col))
        col_idx = column_index_from_string(col)
        inserted = False
        for idx, cell in enumerate(list(row_el)):
            current_ref = cell.attrib.get("r", "")
            current_col = cell_col(current_ref)
            if current_col and column_index_from_string(current_col) > col_idx:
                row_el.insert(idx, new_cell)
                inserted = True
                break
        if not inserted:
            row_el.append(new_cell)
        return new_cell

    def clear_cell_children(self, cell):
        for child in list(cell):
            cell.remove(child)
        for key in ("t", "cm", "vm"):
            cell.attrib.pop(key, None)

    def set_inline_string(self, row_idx, col, value):
        cell = self.get_or_create_cell(row_idx, col)
        self.clear_cell_children(cell)
        if value in (None, ""):
            return
        cell.attrib["t"] = "inlineStr"
        is_el = ET.SubElement(cell, f"{{{MAIN_NS}}}is")
        text_el = ET.SubElement(is_el, f"{{{MAIN_NS}}}t")
        text = str(value)
        text_el.text = text
        if text.strip() != text:
            text_el.attrib["{http://www.w3.org/XML/1998/namespace}space"] = "preserve"

    def set_number(self, row_idx, col, value):
        cell = self.get_or_create_cell(row_idx, col)
        self.clear_cell_children(cell)
        v = ET.SubElement(cell, f"{{{MAIN_NS}}}v")
        v.text = str(value)

    def set_formula(self, cell, formula):
        self.clear_cell_children(cell)
        f = ET.SubElement(cell, f"{{{MAIN_NS}}}f")
        f.text = formula[1:] if formula.startswith("=") else formula

    def formulas_for_row(self, row_idx):
        formulas = {}
        for col_idx in range(1, 23):
            col = get_column_letter(col_idx)
            value = self.formula_ws[f"{col}{row_idx}"].value
            if isinstance(value, str) and value.startswith("="):
                formulas[col] = value
        return formulas

    def translate_formula(self, formula, col, old_row, new_row):
        try:
            return Translator(formula, origin=f"{col}{old_row}").translate_formula(f"{col}{new_row}")
        except Exception:
            return formula.replace(str(old_row), str(new_row))

    def append_match_row(self):
        old_row = self.table_end_row()
        new_row = old_row + 1
        old_el = self.row_elements().get(old_row)
        if old_el is None:
            raise RuntimeError(f"Template row missing: {old_row}")
        formulas = self.formulas_for_row(old_row)
        new_el = copy.deepcopy(old_el)
        new_el.attrib["r"] = str(new_row)
        if new_el.attrib.get("spans"):
            new_el.attrib["spans"] = "1:22"
        for cell in new_el.findall("m:c", NS):
            old_ref = cell.attrib.get("r", "")
            col = cell_col(old_ref)
            if not col:
                continue
            cell.attrib["r"] = f"{col}{new_row}"
            if col in formulas:
                translated = self.translate_formula(formulas[col], col, old_row, new_row)
                self.set_formula(cell, translated)
            elif col in {"C", "D", "E", "F", "V"}:
                self.clear_cell_children(cell)
            elif col not in {"A", "B", "G", "U"}:
                self.clear_cell_children(cell)
        self.sheet_data.append(new_el)

        self.set_number(new_row, "A", new_row - 2)
        self.set_number(new_row, "B", new_row - 2)

        self.update_table_ref(new_row)
        self.update_dimension(new_row)
        return new_row

    def update_table_ref(self, end_row):
        min_col, min_row, max_col, _ = self.table_bounds()
        ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{end_row}"
        self.table_root.attrib["ref"] = ref
        auto_filter = self.table_root.find("m:autoFilter", NS)
        if auto_filter is not None:
            auto_filter.attrib["ref"] = ref

    def update_dimension(self, end_row):
        dimension = self.sheet_root.find("m:dimension", NS)
        if dimension is None:
            return
        ref = dimension.attrib.get("ref", "")
        if ":" not in ref:
            return
        start, end = ref.split(":", 1)
        end_col = cell_col(end) or "V"
        existing_end_row = int(re.sub(r"\D+", "", end) or end_row)
        dimension.attrib["ref"] = f"{start}:{end_col}{max(end_row, existing_end_row)}"

    def ensure_capacity(self, row_idx):
        while self.table_end_row() < row_idx:
            self.append_match_row()

    def set_match_row(self, row_idx, payload):
        self.ensure_capacity(row_idx)
        self.set_inline_string(row_idx, "C", payload["teamA"])
        self.set_inline_string(row_idx, "D", payload["teamB"])
        self.set_inline_string(row_idx, "E", payload["winner"])
        self.set_inline_string(row_idx, "F", payload["resultType"])
        self.set_inline_string(row_idx, "G", payload["tier"])
        self.set_inline_string(row_idx, "U", payload["tournamentName"])
        self.set_inline_string(row_idx, "V", payload["notes"])

    def save(self):
        self.files[self.sheet_path] = ET.tostring(self.sheet_root, encoding="utf-8", xml_declaration=True)
        self.files[self.table_path] = ET.tostring(self.table_root, encoding="utf-8", xml_declaration=True)
        fd, tmp_name = tempfile.mkstemp(prefix=self.path.stem, suffix=".xlsx", dir=str(self.path.parent))
        os.close(fd)
        tmp_path = Path(tmp_name)
        try:
            with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
                for name, data in self.files.items():
                    zout.writestr(name, data)
            tmp_path.replace(self.path)
        finally:
            if tmp_path.exists():
                tmp_path.unlink()


def read_workbook_cached(path):
    return load_workbook(path, data_only=True, keep_links=True)


def team_info(ws, name):
    wanted = normalize_space(name).lower()
    for row in range(3, ws.max_row + 1):
        team = ws.cell(row, 4).value
        if normalize_space(team).lower() == wanted:
            return {
                "name": team,
                "code": ws.cell(row, 5).value,
                "continent": ws.cell(row, 6).value,
                "startingElo": ws.cell(row, 7).value,
                "matchDelta": ws.cell(row, 8).value,
                "bonusPoints": ws.cell(row, 9).value,
                "currentElo": ws.cell(row, 10).value,
                "currentRank": ws.cell(row, 3).value,
                "updatedBy": ws.cell(row, 12).value,
            }
    return None


def previous_matches(ws, name, limit=5):
    wanted = normalize_space(name).lower()
    items = []
    for row in range(2, ws.max_row + 1):
        team = ws.cell(row, 3).value
        if normalize_space(team).lower() != wanted:
            continue
        match_id = ws.cell(row, 2).value
        items.append(
            {
                "matchId": match_id,
                "team": team,
                "opponent": ws.cell(row, 6).value,
                "result": ws.cell(row, 8).value,
                "simplify": ws.cell(row, 9).value,
                "eloChange": ws.cell(row, 10).value,
                "event": ws.cell(row, 11).value,
            }
        )
    items.sort(key=lambda item: item.get("matchId") or 0, reverse=True)
    return items[:limit]


def status():
    path = workbook_path()
    db = load_db(db_path())
    result = {
        "workbook": str(path),
        "exists": path.exists(),
        "syncDb": str(db_path()),
        "lockFileExists": path.with_name(f"~${path.name}").exists(),
    }
    if not path.exists():
        return result
    wb = load_workbook(path, data_only=True, keep_links=True)
    ws = wb["Matches"]
    table_ref = ws.tables["MatchesTable"].ref
    _, start_row, _, end_row = range_boundaries(table_ref)
    empty = 0
    for row in range(start_row + 1, end_row + 1):
        if not ws.cell(row, 3).value and not ws.cell(row, 4).value:
            empty += 1
    result.update({"tableRef": table_ref, "emptySlots": empty})
    return result


def context(payload):
    path = workbook_path()
    if not path.exists():
        raise RuntimeError(f"Workbook not found: {path}")
    wb = read_workbook_cached(path)
    teams_ws = wb["Teams"]
    hist_ws = wb["Hist"]
    names = [payload.get("teamA"), payload.get("teamB")]
    return {
        "workbook": str(path),
        "teams": {name: team_info(teams_ws, name) for name in names if name},
        "history": {name: previous_matches(hist_ws, name) for name in names if name},
        "lockFileExists": path.with_name(f"~${path.name}").exists(),
    }


def notes_with_code(payload, updated_at):
    score = payload.get("score")
    score_text = f" | score {score}" if score else ""
    return f"[TCODE:{payload['matchCode']}]{score_text} | submitted {updated_at} | source Tourney"


def scan_row_by_code(ws, match_code):
    marker = f"[TCODE:{match_code}]"
    for row in range(3, ws.max_row + 1):
        note = ws.cell(row, 22).value
        if isinstance(note, str) and marker in note:
            return row
    return None


def first_blank_row(ws):
    _, start_row, _, end_row = range_boundaries(ws.tables["MatchesTable"].ref)
    for row in range(start_row + 1, end_row + 1):
        if not ws.cell(row, 3).value and not ws.cell(row, 4).value:
            return row
    return end_row + 1


def normalize_submit_payload(payload):
    required = ["matchCode", "tournamentName", "teamA", "teamB", "winner", "resultType", "tier"]
    missing = [key for key in required if not normalize_space(payload.get(key))]
    if missing:
        raise RuntimeError(f"Missing submit fields: {', '.join(missing)}")
    winner = payload["winner"]
    if winner not in {"Team A", "Team B", "Tie"}:
        raise RuntimeError("Winner must be Team A, Team B, or Tie")
    if payload["resultType"] not in {"Hoshin-Tora", "Hoshin-Kai", "Hoshin-Renga", "Renga"}:
        raise RuntimeError("Invalid Kitakana result type")
    if not re.match(r"^Tier [1-5]$", payload["tier"]):
        raise RuntimeError("Invalid tier")
    return {**payload, "matchCode": normalize_space(payload["matchCode"])}


def submit_one(payload, db, editor=None, cached_wb=None):
    payload = normalize_submit_payload(payload)
    path = workbook_path()
    if not path.exists():
        raise RuntimeError(f"Workbook not found: {path}")
    updated_at = now_iso()
    payload["notes"] = notes_with_code(payload, updated_at)

    wb = cached_wb or load_workbook(path, data_only=False, keep_links=True)
    ws = wb["Matches"]
    record = db.setdefault("matches", {}).get(payload["matchCode"])
    row = record.get("excelRow") if record else None
    if not row:
        row = scan_row_by_code(ws, payload["matchCode"])
    if not row:
        row = first_blank_row(ws)

    owns_editor = editor is None
    if owns_editor:
        make_backup(path, db)
        editor = XlsxXmlEditor(path)
    editor.set_match_row(row, payload)
    if owns_editor:
        editor.save()

    db["matches"][payload["matchCode"]] = {
        "excelRow": row,
        "updatedAt": updated_at,
        "tournamentName": payload["tournamentName"],
        "teamA": payload["teamA"],
        "teamB": payload["teamB"],
        "winner": payload["winner"],
        "resultType": payload["resultType"],
        "tier": payload["tier"],
        "sourceMatchId": payload.get("sourceMatchId"),
        "score": payload.get("score"),
    }
    return {"matchCode": payload["matchCode"], "excelRow": row, "updatedAt": updated_at}


def submit(payload):
    path = workbook_path()
    db = load_db(db_path())
    make_backup(path, db)
    result = submit_one(payload, db)
    save_db(db_path(), db)
    return result


def batch(payload):
    matches = payload.get("matches") or []
    if not isinstance(matches, list) or not matches:
        raise RuntimeError("No matches supplied")
    path = workbook_path()
    db = load_db(db_path())
    make_backup(path, db)
    cached_wb = load_workbook(path, data_only=False, keep_links=True)
    editor = XlsxXmlEditor(path)
    results = []
    errors = []
    seen = set()
    for item in matches:
        code = normalize_space(item.get("matchCode"))
        if not code or code in seen:
            continue
        seen.add(code)
        try:
            results.append(submit_one(item, db, editor=editor, cached_wb=cached_wb))
        except Exception as exc:
            errors.append({"matchCode": code, "error": str(exc)})
    editor.save()
    save_db(db_path(), db)
    return {"submitted": len(results), "results": results, "errors": errors}


def main():
    raw = sys.stdin.read()
    payload = json.loads(raw or "{}")
    op = payload.get("op")
    try:
        if op == "status":
            result = status()
        elif op == "context":
            result = context(payload)
        elif op == "submit":
            result = submit(payload)
        elif op == "batch":
            result = batch(payload)
        else:
            raise RuntimeError(f"Unknown operation: {op}")
        json_out({"ok": True, **result})
    except Exception as exc:
        json_out({"ok": False, "error": str(exc)})


if __name__ == "__main__":
    main()
